from io import BytesIO
from typing import List

import pandas as pd
from openpyxl.cell import WriteOnlyCell
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from app.core.constants import Direction, Role
from app.schemas.ExternalTransactionScheme import (
    ExportSumTransactionsResponse,
    ExportTransactionsResponse,
)
from app.utils.measure import measure_exe_time


def adjust_column_widths(ws, max_lengths):
    for col_idx, max_length in max_lengths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max_length + 2


@measure_exe_time
async def generate_export_transactions_xls_report(
    transactions: List[ExportTransactionsResponse], role: str
):
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="Transactions Report")

    headers = [
        "Transaction ID",
        "Create Timestamp",
        "Direction",
        "Bank Detail Number",
        "Transaction Amount",
        "Status",
        "USDT Deposit Change",
        "Profit Change",
        "Exchange Rate",
        "Fiat Balance Change",
        "Interest Rate",
    ]
    ws.append(headers)

    max_lengths = {i: len(header) for i, header in enumerate(headers, start=1)}
    rows = []

    for exported in transactions:
        profit = (
            exported.profit_change
            if exported.direction == Direction.INBOUND
            else exported.usdt_deposit_change
            - exported.transaction_amount / exported.exchange_rate
        )
        row = [
            exported.transaction_id,
            exported.create_timestamp.strftime("%Y-%m-%d %H:%M:%S"),
            exported.direction,
            exported.bank_detail_number,
            exported.transaction_amount,
            exported.status,
            exported.usdt_deposit_change,
            profit if role != Role.MERCHANT else 0,
            exported.exchange_rate,
            exported.fiat_balance_change,
            exported.interest,
        ]
        write_only_row = [WriteOnlyCell(ws, value=item) for item in row]
        rows.append(write_only_row)

        for col_idx, cell in enumerate(write_only_row, start=1):
            max_lengths[col_idx] = max(
                max_lengths.get(col_idx, 0),
                len(str(cell.value) if cell.value is not None else ""),
            )

    for row in rows:
        ws.append(row)

    adjust_column_widths(ws, max_lengths)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


async def export_transactions_to_csv(
    transactions: List[ExportSumTransactionsResponse],
) -> bytes:
    data = [transaction.dict() for transaction in transactions]

    df = pd.DataFrame(data)

    csv_data = df.to_csv(index=False)

    return csv_data.encode("utf-8")
