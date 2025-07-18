import asyncio
from datetime import datetime, timedelta
from io import BytesIO
from itertools import chain

from openpyxl import Workbook
from openpyxl.styles import numbers, PatternFill, Border, Side
from sqlalchemy import text, select, and_, true

from app.core.constants import Direction, Status
from app.core.session import async_session
from app.core import redis
from app.models import BankDetailModel, UserModel, TeamModel, MerchantModel


async def get_daily_traffic_stats(merchant_ids: list[str],
                                  agent_id: str,
                                  day_period=7,
                                  hour_start=7,
                                  is_daily=True,
                                  ) -> bytes:
    single_quote = "'"
    stringed_list = f"{','.join([single_quote + str(i) + single_quote for i in merchant_ids])}"
    
    wk = Workbook()
    wb = wk.active
    wb['A1'], wb['B1'] = 'Дата от (UTC)', 'Дата до (UTC)'
    wb['C1'], wb['D1'], wb['E1'], wb['F1'], wb['G1'], wb['H1'], wb['I1'], wb['J1'], wb['K1'], wb['L1'], wb['M1'], wb['N1'] = (
        'inbound (accept)', 'outbound (accept)',
        'inbound (all)', 'outbound (all)',
        'Count inbound (all)', 'Count inbound (success)',
        'Count outbound (all)', 'Count outbound (success)',
        'count pay in conversion',
        'count pay out conversion',
        'прибыль платформы (USDT)',
        'вывод прибыли платформы (USDT)'
    )
    if is_daily:
        today = datetime.utcnow().replace(hour=hour_start, minute=0, second=0, microsecond=0)
    else:
        today = datetime.utcnow().replace(minute=0, second=0, microsecond=0)
    for days in range(0, day_period):
        if is_daily:
            date_to = today - timedelta(days=days)
            date_from = today - timedelta(days=days + 1)
        else:
            date_to = today - timedelta(hours=days)
            date_from = today - timedelta(hours=days + 1)
        async with async_session() as session:
            query = await session.execute(
                text(f"""select
        sum(case when direction='{Direction.INBOUND}' and status='{Status.ACCEPT}'
         then amount else 0 end) / 1000000 "inbound_accept",
        sum(case when direction='{Direction.OUTBOUND}' and status='{Status.ACCEPT}'
         then amount else 0 end) / 1000000 "outbound_accept",
        sum(case when direction='{Direction.INBOUND}'
         then amount else 0 end) / 1000000 "inbound_all",
        sum(case when direction='{Direction.OUTBOUND}'
         then amount else 0 end) / 1000000 "outbound_all",
         
        sum(case when direction='{Direction.INBOUND}'
         then 1 else 0 end) "count_inbound_all",
        sum(case when direction='{Direction.INBOUND}' and status='{Status.ACCEPT}'
         then 1 else 0 end) "count_inbound_accept",
        sum(case when direction='{Direction.OUTBOUND}'
         then 1 else 0 end) "count_inbound_all",
        sum(case when direction='{Direction.OUTBOUND}' and status='{Status.ACCEPT}'
         then 1 else 0 end) "count_outbound_accept",
         
        1.0 * sum(case when direction='{Direction.INBOUND}' and status='{Status.ACCEPT}'
         then 1 else 0 end) / (1.0 * greatest(cast(1 as BIGINT), sum(case when direction='{Direction.INBOUND}'
         then 1 else 0 end))) "count_inbound_conversion",
         
        1.0 * sum(case when direction='{Direction.OUTBOUND}' and status='{Status.ACCEPT}'
         then 1 else 0 end) / (1.0 * greatest(cast(1 as BIGINT), sum(case when direction='{Direction.OUTBOUND}'
         then 1 else 0 end))) "count_outbound_conversion"
         
                from external_transaction_model
                where merchant_id in (
                    {stringed_list}
                )
                  AND create_timestamp < '{date_to}'
                  AND create_timestamp >= '{date_from}'
                """)
            )
            
            for name, i in zip(['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L'],
                               query.first()):
                wb[f'{name}{days + 2}'].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                if name in ['G', 'H', 'I', 'J'] and i is not None:
                    wb[f'{name}{days + 2}'] = int(i)
                else:
                    wb[f'{name}{days + 2}'] = i
                
                wb[f'A{days + 2}'] = date_from
                wb[f'B{days + 2}'] = date_to
        async with async_session() as session:
            query = await session.execute(
                text(f"""
                select
                sum(case when (profit_balance > 0 or trust_balance > 0)
                 then profit_balance + trust_balance + locked_balance else 0 end) / 1000000,
                sum(case when (profit_balance < -1000000000 or trust_balance < -1000000000)
                 then profit_balance + trust_balance + locked_balance else 0 end) / 1000000
                from user_balance_change_model
                where user_id = '{agent_id}'
                  AND create_timestamp < '{date_to}'
                  AND create_timestamp >= '{date_from}'
                """)
            )
            for name, i in zip(['M', 'N'], query.first()):
                wb[f'{name}{days + 2}'] = i
    pretty(wb, scale=4)
    output = BytesIO()
    wk.save(output)
    return output.getvalue()


def fill_mx_mn(wb):
    mx = dict()
    mn = dict()
    for cell in chain.from_iterable(wb.iter_cols()):
        try:
            float(cell.value)
        except (ValueError, TypeError):
            continue
        if cell.column_letter not in mn.keys():
            mn[cell.column_letter] = float(cell.value)
        mn[cell.column_letter] = min(mn[cell.column_letter], float(cell.value))
        if cell.column_letter not in mx.keys():
            mx[cell.column_letter] = float(cell.value)
        mx[cell.column_letter] = max(mx[cell.column_letter], float(cell.value))
    
    return mx, mn


def pretty(wb, scale=2):
    mx, mn = fill_mx_mn(wb)
    for cell in chain.from_iterable(wb.iter_cols()):
        if str(cell.value):
            wb.column_dimensions[cell.column_letter].width = max(
                wb.column_dimensions[cell.column_letter].width,
                len(f"{cell.value}"),
            )
            if cell.column_letter in mx and cell.column_letter in mn and mx[cell.column_letter] != mn[cell.column_letter]:
                try:
                    delta = int(255 - (float(cell.value) - mn[cell.column_letter])
                                / (mx[cell.column_letter] - mn[cell.column_letter]) * 255) // scale
                    hex_color = f"{255:02X}{255 - delta:02X}{255 - delta:02X}"
                    cell.fill = PatternFill(start_color=hex_color,
                                            end_color=hex_color,
                                            fill_type='solid'
                                            )
                except (ValueError, TypeError):
                    pass
            cell.border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))


async def get_conversion_stats(merchant_ids: list[str],
                               hour_period=1,
                               ) -> bytes:
    single_quote = "'"
    stringed_list = f"{','.join([single_quote + str(i) + single_quote for i in merchant_ids])}"
    wk = Workbook()
    wb = wk.active
    wb['A1'], wb['B1'], wb['C1'] = ('Название команды',
                                    f'К-во pay in за {hour_period}h',
                                    f'К-во pay out за {hour_period}h')
    wb['D1'], wb['E1'], wb['F1'] = (
        f'Конверсия pay in за {hour_period}h',
        f'Конверсия pay out за {hour_period}h',
        f'Общая конверсия за {hour_period}h'
    )
    date_from = datetime.utcnow() - timedelta(hours=hour_period)
    print(date_from)
    async with async_session() as session:
        query = await session.execute(
            text(f"""
            
            SELECT name,
                   sum(case when direction='inbound' then 1 else 0 end) "inbound_all",
                   sum(case when direction='outbound' then 1 else 0 end) "outbound_all",
                   1.0 * sum(case when status='accept' and direction='inbound' then 1 else 0 end)
                             / (abs(sum(case when direction='inbound' then 1 else 0 end ) - 1) + 1) conv_in,
                   1.0 * sum(case when status='accept' and direction='outbound' then 1 else 0 end)
                             / (abs(sum(case when direction='outbound' then 1 else 0 end ) - 1) + 1) conv_out,
                   1.0 * sum(case when status='accept' then 1 else 0 end)
                             / (abs(count(*) - 1) + 1) conv_total
            FROM external_transaction_model e inner join traffic_weight_contact_model t
                                                         on e.team_id=t.team_id and e.merchant_id = t.merchant_id
                                              inner join user_model u on u.id = e.team_id
                AND e.merchant_id in (
                    {stringed_list}
                )
                AND e.create_timestamp >= '{date_from}'
            GROUP BY name, u.id order by inbound_all desc;
            """)
        )
        last_it = 0
        for it, q in enumerate(query.fetchall()):
            for name, i in zip(['A', 'B', 'C', 'D', 'E', 'F'],
                               q):
                if name in ['D', 'E', 'F']:
                    wb[f'{name}{it + 2}'].number_format = '#,##0.00%'
                else:
                    wb[f'{name}{it + 2}'].number_format = '#,##0'
                wb[f'{name}{it + 2}'] = i
            last_it = it

        pretty(wb, scale=2)

        namespace_subquery = (await session.execute(
            select(MerchantModel.namespace_id)
            .distinct()
            .where(MerchantModel.id.in_(merchant_ids))
        )).subquery()

        types = (await session.execute(
            select(BankDetailModel.type)
            .distinct()
            .join(TeamModel, BankDetailModel.team_id == TeamModel.id)
            .where(
                and_(
                    BankDetailModel.is_active == true(),
                    TeamModel.namespace_id.in_(namespace_subquery)
                )
            )
            .order_by(BankDetailModel.type)
        ))

        count = {}

        for e in types:
            count[e] = 0

        wb[f'A{last_it + 5}'] = "count_450_1h"

        for merch in merchant_ids:
            for e in types:
                count[e] += await get_450errors_count(merch, e, date_from=date_from, date_to=datetime.utcnow())

        letter_ind = 2

        for k, v in count.items():
            wb.cell(row=last_it + 4, column=letter_ind, value=k)
            wb.cell(row=last_it + 5, column=letter_ind, value=v)
            letter_ind += 1

        wb[f'A{last_it + 6}'] = "total pay in"

        qry = await session.execute(text(f"""
                SELECT e.type, COUNT(*) as count
                FROM external_transaction_model e
                WHERE e.create_timestamp >= '{date_from}' AND e.merchant_id IN ({stringed_list}) AND e.direction = 'inbound'
                GROUP BY e.type ORDER BY e.type;
            """))
        
        rows = qry.fetchall()

        result_count = {row.type: row.count for row in rows}

        print(result_count)

        letter_ind = 2

        for t in types:
            if t in result_count:
                wb.cell(row=last_it + 6, column=letter_ind, value=result_count[t])
            else:
                wb.cell(row=last_it + 6, column=letter_ind, value=0)
            letter_ind += 1
        
    output = BytesIO()
    wk.save(output)
    return output.getvalue()


async def get_teams_info(merchant_ids: list[str],
                         ) -> bytes:
    single_quote = "'"
    stringed_list = f"{','.join([single_quote + str(i) + single_quote for i in merchant_ids])}"

    wk = Workbook()
    wb = wk.active
    (wb['A1'], wb['B1'], wb['C1'],
     wb['D1'], wb['E1'], wb['F1'], wb['G1'], wb['H1'], wb['I1']
     ) = ('Название команды',
          f'Card profiles',
          f'Phone profiles',
          f'Cards',
          f'Phones',
          f'Balances (USDT)',
          f'Limit (USDT)',
          f'Pending pay in',
          f'Pending pay out'
          )
    async with async_session() as session:
        query = await session.execute(
            text(f"""
            
            SELECT Q.name, N.card_profiles, N.phone_profiles, G.card, G.phone, Q.debt, Q.credit_factor, PT.pi, OT.por

FROM (SELECT T.name,
             sum(case when T.card > 0 then 1 else 0 end)  card_profiles,
             sum(case when T.phone > 0 then 1 else 0 end) phone_profiles
      FROM (SELECT u.name,
                   sum(case
                           when u.is_inbound_enabled and b.is_active and not b.is_deleted
                               and b.type = 'card' then 1
                           else 0 end) "card",
                   sum(case
                           when u.is_inbound_enabled
                               and b.is_active and not b.is_deleted and b.type = 'phone' then 1
                           else 0 end) "phone"
            FROM bank_detail_model b inner join
                     (SELECT DISTINCT uu.id,
                                                 name,
                                                 is_inbound_enabled,
                                                 is_outbound_enabled,
                                                 team_id,
                                                 credit_factor
                                 FROM teams uu
                                          inner join traffic_weight_contact_model t
                                                     on uu.id = t.team_id
                                                         and t.merchant_id in (
                                                                               {stringed_list}
                                                             ) AND
                                                        (t.outbound_traffic_weight > 0 or t.inbound_traffic_weight > 0)) u
                                on u.team_id = b.team_id
            GROUP BY u.name, u.id, b.device_hash
            order by card desc) T
      group by T.name) N
         RIGHT join
     (SELECT u.name,
             sum(case
                     when u.is_inbound_enabled and b.is_active and not b.is_deleted
                         and b.type = 'card' then 1
                     else 0 end) "card",
             sum(case
                     when u.is_inbound_enabled
                         and b.is_active and b.type = 'phone' and not b.is_deleted then 1
                     else 0 end) "phone"
      FROM bank_detail_model b inner join
               (SELECT DISTINCT uu.id,
                                                 name,
                                                 is_inbound_enabled,
                                                 is_outbound_enabled,
                                                 team_id,
                                                 credit_factor
                                 FROM teams uu
                                          inner join traffic_weight_contact_model t
                                                     on uu.id = t.team_id
                                                         and t.merchant_id in (
                                                                              {stringed_list}
                                                             ) AND
                                                        (t.outbound_traffic_weight > 0 or t.inbound_traffic_weight > 0)) u
                                on u.team_id = b.team_id
      GROUP BY u.name, u.id
      order by card desc) G on N.name = G.name
         RIGHT JOIN
     (SELECT u.name, u.credit_factor, sum(c.trust_balance + c.profit_balance + c.locked_balance) / 1000000 debt
      FROM user_balance_change_model c
               inner join (SELECT DISTINCT name, balance_id, team_id, credit_factor FROM teams uu
               inner join traffic_weight_contact_model t
                          on uu.id = t.team_id

                          and t.merchant_id in (
                                {stringed_list}
                                  ) AND (t.outbound_traffic_weight > 0 or t.inbound_traffic_weight > 0)) u
                          on u.balance_id = c.balance_id


      GROUP BY u.name, u.credit_factor, u.balance_id) Q on Q.name = N.name
      LEFT JOIN (
        SELECT name, count(*) "pi"
        FROM external_transaction_model t inner join teams u on t.team_id = u.id
        WHERE status = 'pending'
        AND direction = 'inbound'
        AND merchant_id in ({stringed_list})
        GROUP BY name
      ) PT on PT.name = Q.name
      LEFT JOIN (
        SELECT name, count(*) "por"
        FROM external_transaction_model t inner join teams u on t.team_id = u.id
        WHERE status = 'pending'
        AND direction = 'outbound'
        AND merchant_id in ({stringed_list})
        GROUP BY name
      ) OT on OT.name = Q.name

order by card_profiles desc;
            
            """)
        )
        part1 = []
        part2 = []
        for it, q in enumerate(query.fetchall()):
            isGreen = False
            for name, i in zip(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I'],
                               q):
                if name in ['F']:
                    wb[f'{name}{it + 2}'].number_format = '#,##0.00'
                elif name in ['B', 'C', 'D', 'E', 'G', 'H', 'I']:
                    wb[f'{name}{it + 2}'].number_format = '#,##0'
                if name in ['A']:
                    result = await session.execute(text(
                        "SELECT id FROM user_model WHERE name = :name"),
                        {"name": i}
                    )
                    row = result.fetchone()
                    user_id = row[0]
                    res = await session.execute(text("""SELECT * 
                                            FROM traffic_weight_contact_model 
                                            WHERE (inbound_traffic_weight > 0) AND (team_id = :user_id) AND merchant_id = ANY(
                                                    :merchant_ids
                                            )
                                            """), {"user_id": user_id, "merchant_ids": merchant_ids}
                    )
                    rows = res.fetchall()
                    if len(rows) > 0:
                        isGreen = True
                    else:
                        isGreen = False
                if isGreen == True:
                    part1.append([name, i])
                else:
                    part2.append([name, i])
        it = 0
        for name, i in part1:
            wb[f'{name}{it + 2}'] = i
            if name in ['A']:
                hex_color = "00ff00"
                wb[f'{name}{it + 2}'].fill = PatternFill(start_color=hex_color,
                                                         end_color=hex_color,
                                                         fill_type='solid'
                                                         )
            if name in ['I']:
                it += 1
        for name, i in part2:
            wb[f'{name}{it + 2}'] = i
            if name in ['I']:
                it += 1
        pretty(wb, scale=5)
    output = BytesIO()
    wk.save(output)
    return output.getvalue()

async def get_450errors_count(
    merchant_id: str,
    type: str | None,
    date_from: datetime,
    date_to: datetime,
    bank: str | None = None,
    payment_system: str | None = None,
    is_vip: str | None = None
) -> int:
    date_from_ts = int(date_from.timestamp())
    date_to_ts = int(date_to.timestamp())

    type = type or 'None'
    bank = bank or 'None'
    payment_system = payment_system or 'None'

    is_vip_options = [is_vip] if is_vip is not None else ["true", "false"]

    keys = await redis.rediss.smembers(f"/count/errors/450/index/{merchant_id}")
    if not keys:
        return 0

    filtered_keys = []
    for key in keys:
        key_str = key.decode()
        *_, k_type, k_bank, k_payment, k_is_vip = key_str.split("/")
        if k_type != type:
            continue
        if bank != 'None' and k_bank != bank:
            continue
        if payment_system != 'None' and k_payment != payment_system:
            continue
        if k_is_vip not in is_vip_options:
            continue
        filtered_keys.append(key)

    if len(filtered_keys) == 0:
        return 0

    counts = await asyncio.gather(*[
        redis.rediss.zcount(key, date_from_ts, date_to_ts)
        for key in filtered_keys
    ])
    return sum(counts)





if __name__ == '__main__':
    asyncio.run(get_teams_info(merchant_ids=['a5a25687-ab15-4b98-bc37-dc0a1107f888'],
                                 ))
# asyncio.run(get_daily_traffic_stats(merchant_id='a5a25687-ab15-4b98-bc37-dc0a1107f888',
#                                     agent_id='c46b2a87-1424-40aa-b1da-5c8cf696a2f7'
#                                     ))
# asyncio.run(get_teams_info(merchant_id='a5a25687-ab15-4b98-bc37-dc0a1107f888',
#                            ))
